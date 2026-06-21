from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
import os

# ==========================================
# НАСТРОЙКИ
# ==========================================

TEMPLATE = "глянц 3.png"
OUTPUT_FOLDER = "output"

FONT_FILE = "Montserrat-SemiBold.ttf"

TEXT_COLOR = (20, 30, 80)

# Шаблоны для обработки
TEMPLATES = [
    ("глянц 3.png", "gloss_3", "глянц", "3шт"),
    ("глянц 2.png", "gloss_2", "глянц", "2шт"),
    ("антишпион 2.png", "antispion_2", "антишпион", "2шт"),
    ("антишпион 3.png", "antispion_3", "антишпион", "3шт"),
]

# область текста
TEXT_BOX = (
    70,   # left
    165,  # top
    950,  # right
    270   # bottom
)

# ==========================================


def load_models_from_excel(xlsx_file):
    wb = load_workbook(xlsx_file)
    ws = wb.active

    models = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        value = row[0]

        if value:
            models.append(str(value).strip())

    return models


def create_image(model_name, template_path, template_name, glass_type, quantity):

    img = Image.open(template_path).convert("RGBA")

    draw = ImageDraw.Draw(img)

    font_size = 120

    left, top, right, bottom = TEXT_BOX
    box_width = right - left
    box_height = bottom - top

    while font_size > 40:
        font = ImageFont.truetype(FONT_FILE, font_size)

        bbox = draw.textbbox((0, 0), model_name, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]

        if text_width <= box_width and text_height <= box_height:
            break

        font_size -= 2

    bbox = draw.textbbox((0, 0), model_name, font=font)

    text_width = bbox[2] - bbox[0]
    text_height = bbox[3] - bbox[1]

    bottom_margin = 15  # отступ от нижней границы рамки
    text_x = left + (box_width - text_width) // 2
    text_y = bottom - text_height - bottom_margin

    draw.text(
        (text_x, text_y),
        model_name,
        fill=(17, 31, 95),
        font=font
    )

    safe_name = (
        model_name
        .replace("/", "-")
        .replace("\\", "-")
        .replace(":", "-")
        .replace("*", "")
        .replace("?", "")
        .replace('"', "")
        .replace("<", "")
        .replace(">", "")
        .replace("|", "")
    )

    # Добавляем информацию о типе стекла и количестве в имя файла
    safe_name = f"{safe_name} - {glass_type} {quantity}"

    output_folder = os.path.join(OUTPUT_FOLDER, template_name)
    os.makedirs(output_folder, exist_ok=True)

    output_file = os.path.join(
        output_folder,
        f"{safe_name}.png"
    )

    img.save(output_file)

    print(f"Создано: {output_file}")


from models import GLOSS_AND_ANTI, GLOSS_ONLY

def main():

    models = GLOSS_AND_ANTI + GLOSS_ONLY

    print(f"Найдено моделей: {len(models)}\n")

    for template_path, template_name, glass_type, quantity in TEMPLATES:
        print(f"Обработка {template_name}...")
        for model in models:
            create_image(model, template_path, template_name, glass_type, quantity)
        print(f"[OK] {template_name} готово\n")

    print("Все готово.")

if __name__ == "__main__":
    main()